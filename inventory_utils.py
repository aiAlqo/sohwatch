import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font
from io import BytesIO
from typing import List, Dict, Any, Optional

def get_row_fill_color(status: str) -> str:
    """
    Map inventory status to a hex color code for row highlighting.
    """
    return {
        "🔴 Critical!!! Below Min Qty": "#FFCCCC",
        "🟠 Reorder Level": "#FFE4B3",
        "🕣 Overstocked": "#FFCCFF",
        "✅ Healthy": "#CCFFCC",
        "❓ Missing SOH": "#E0E0E0"
    }.get(status, "#FFFFFF")

def generate_excel(df_to_export: pd.DataFrame) -> BytesIO:
    """
    Export a DataFrame to an Excel file with color-coded rows based on status.
    Freezes the header row and auto-adjusts column widths.
    """
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Status"
    headers = df_to_export.columns.tolist()
    ws.append(headers)
    for _, row in df_to_export.iterrows():
        row_values = row.tolist()
        ws.append(row_values)
        status = row.get("Status", "")
        fill_color = get_row_fill_color(status)
        if fill_color:
            for col in range(1, len(row_values) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.fill = PatternFill(start_color=fill_color.replace('#', ''),
                                        end_color=fill_color.replace('#', ''),
                                        fill_type="solid")
                cell.font = Font(color="000000")
    # Freeze header row
    ws.freeze_panes = ws['A2']
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(output)
    output.seek(0)
    return output

def assess_status(row: pd.Series) -> str:
    """
    Assess inventory status based on SOH, Min Qty, and Max Qty.
    Handles missing or NaN values robustly.
    """
    soh = row.get('SOH', np.nan)
    min_qty = row.get('Min Qty', np.nan)
    max_qty = row.get('Max Qty', np.nan)
    if pd.isna(soh):
        return "❓ Missing SOH"
    if pd.isna(min_qty) or pd.isna(max_qty):
        return "❓ Missing SOH"
    reorder_threshold = max_qty - ((max_qty - min_qty) / 3)
    if soh < min_qty:
        return "🔴 Critical!!! Below Min Qty"
    elif soh < reorder_threshold:
        return "🟠 Reorder Level"
    elif soh > max_qty:
        return "🕣 Overstocked"
    else:
        return "✅ Healthy"

def suggest_reorder(row: pd.Series) -> Optional[int]:
    """
    Suggest reorder quantity if SOH is below the reorder threshold (i.e., status is 'Reorder Level' or 'Critical!!! Below Min Qty'),
    rounding up to the nearest Minor Order Multiple. Returns None if no reorder is needed or if required fields are missing.
    """
    soh = row.get('SOH', np.nan)
    min_qty = row.get('Min Qty', np.nan)
    max_qty = row.get('Max Qty', np.nan)
    moq = row.get('MOQ', np.nan)
    minor_mult = row.get('Minor Order Multiple', np.nan)
    if pd.isna(soh) or pd.isna(min_qty) or pd.isna(max_qty) or pd.isna(moq):
        return None
    reorder_threshold = max_qty - ((max_qty - min_qty) / 3)
    if soh < reorder_threshold:
        base = max(moq, min_qty - soh)
        if not pd.isna(minor_mult) and minor_mult > 0:
            base = (np.ceil(base / minor_mult)) * minor_mult
        return int(base)
    return None

def highlight_row(row: pd.Series) -> List[str]:
    """
    Return a list of style strings for DataFrame row styling in Streamlit.
    """
    status = row.get("Status", "")
    fill_color = get_row_fill_color(status)
    return [f"background-color: {fill_color}; color: black;" for _ in row]

def simulate_runout(row: pd.Series, forecast_cols: List[str]) -> pd.Series:
    """
    Simulate inventory runout over forecast periods, marking '✅' if SOH covers usage.
    Returns a Series with the same index as forecast_cols.
    """
    remaining_soh = row.get("SOH", np.nan)
    result = []
    for col in forecast_cols:
        usage = row.get(col, np.nan)
        if pd.isna(usage) or pd.isna(remaining_soh):
            result.append("")
        elif remaining_soh >= usage:
            remaining_soh -= usage
            result.append("✅")
        else:
            result.append("")
    return pd.Series(result, index=forecast_cols)

def highlight_forecast(val: Any) -> str:
    """
    Highlights cells green if value is '✅'.
    """
    return 'background-color: lightgreen' if val == "✅" else '' 
